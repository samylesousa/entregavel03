import openpyxl
from elementos import Tarefas, Recursos
from heranca_prioridade import heranca_prioridade

wb = openpyxl.load_workbook('tabela1.xlsx')
ws = wb.active


teto_prioridade = {}
tarefas_recursos = {}

for row in ws.iter_rows(values_only=True):
    print(row)

#organizando os recursos e quais tarefas usam quais recursos
for i in range(2, ws.max_row+1):
    valores_recursos = ws.cell(row=i, column=5).value

    if isinstance(valores_recursos, str):
        if valores_recursos.startswith('[') and valores_recursos.endswith(']'):
            valores_recursos = valores_recursos[1:-1].replace(" ", "").split(",")

    tarefas_recursos[ws.cell(row=i, column=1).value] = valores_recursos


#calculando o teto de prioridade de cada recurso baseado nos dados da tabela
#COLUNA 6
for i in range(2, ws.max_row+1):
    #verificando se é um conjunto de recursos ou apenas 1 recurso
    if len(ws.cell(row=i, column=5).value)<3:
        if ws.cell(row=i, column=5).value not in teto_prioridade:
            teto_prioridade[ws.cell(row=i, column=5).value] = ws.cell(row=i, column=6).value
        elif ws.cell(row=i, column=5).value in teto_prioridade and ws.cell(row=i, column=6).value>teto_prioridade[ws.cell(row=i, column=5).value]:
            teto_prioridade[ws.cell(row=i, column=5).value] = ws.cell(row=i, column=6).value
    else:
        recursos_prioridades = ws.cell(row=i, column=5).value[1:-1].replace(" ", "").split(",")
        print(recursos_prioridades)
        for j in range(0, len(recursos_prioridades)):
            if recursos_prioridades[j] not in teto_prioridade:
                teto_prioridade[recursos_prioridades[j]] = ws.cell(row=i, column=6).value
            elif recursos_prioridades[j] in teto_prioridade and ws.cell(row=i, column=6).value>teto_prioridade[recursos_prioridades[j]]:
                teto_prioridade[recursos_prioridades[j]] = ws.cell(row=i, column=6).value


recursos = []

#criando os recursos
for i in range(0, len(teto_prioridade)):
    recursos.append(
        Recursos(
            list(teto_prioridade.keys())[i],
            teto_prioridade[list(teto_prioridade.keys())[i]]
        )
    )


tarefas = []

#criando as Tarefas de acordo com os elementos que estão no arquivo excel
for i in range(2, ws.max_row+1):
    tarefas.append(
        Tarefas(
            ws.cell(row=i, column=1).value, 
            ws.cell(row=i, column=2).value, 
            ws.cell(row=i, column=4).value, 
            ws.cell(row=i, column=3).value, 
            tarefas_recursos[ws.cell(row=i, column=1).value],
            ws.cell(row=i, column=6).value
        )
    )



# order = heranca_prioridade(tarefas, recursos)
# print(order)

heranca_prioridade(tarefas, recursos, teto_prioridade)

